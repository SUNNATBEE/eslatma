"""
handlers/admin_extras.py — Statistika, Excel eksport, Broadcast,
                           O'quvchi qo'shish, Vaqt sozlash, Faollik.
"""
import io
import logging
from datetime import datetime, timedelta

from aiogram import Bot, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, Message, InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder

from config import ADMIN_IDS, SEND_HOUR, SEND_MINUTE
from credentials import MARS_GROUPS
from database import AudienceType, DatabaseService
from keyboards import kb_back_to_panel, kb_hw_groups

logger = logging.getLogger(__name__)
router = Router()


# ════════════════════════════════════════════════════════════════════════════════
# STATISTIKA
# ════════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data == "admin:stats")
async def admin_stats(cb: CallbackQuery, db: DatabaseService) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    students  = await db.get_all_students()
    today_str = datetime.now().strftime("%Y-%m-%d")
    att_today = await db.get_attendance_by_date(today_str)

    yes_ids  = {a.user_id for a in att_today if a.status == "yes"}
    no_ids   = {a.user_id for a in att_today if a.status == "no"}
    answered = len(yes_ids) + len(no_ids)

    # Per-group breakdown
    from collections import Counter
    group_count = Counter(s.group_name for s in students)
    group_lines = "\n".join(f"  • {g}: {group_count.get(g, 0)} ta" for g in MARS_GROUPS)

    text = (
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Jami ro'yxatdan o'tganlar: <b>{len(students)}</b> ta\n\n"
        f"📚 Guruhlar bo'yicha:\n{group_lines}\n\n"
        f"📅 Bugungi davomat ({today_str}):\n"
        f"  ✅ Boraman: <b>{len(yes_ids)}</b> ta\n"
        f"  ❌ Kela olmaydi: <b>{len(no_ids)}</b> ta\n"
        f"  ⏳ Javob bermagan: <b>{len(students) - answered}</b> ta"
    )

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Panel", callback_data="admin:panel"))
    try:
        await cb.message.edit_text(text, reply_markup=builder.as_markup())
    except TelegramBadRequest:
        await cb.message.answer(text, reply_markup=builder.as_markup())
    await cb.answer()


# ════════════════════════════════════════════════════════════════════════════════
# EXCEL EKSPORT
# ════════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data == "admin:excel_export")
async def admin_excel_export(cb: CallbackQuery, db: DatabaseService, bot: Bot) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    await cb.answer("⏳ Tayyorlanmoqda...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        students = await db.get_all_students()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"

        # Header
        headers = ["#", "Ism Familya", "Guruh", "Mars ID", "Telefon", "Telegram", "Ro'yxatdan"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 18

        # Data (sorted by group then name)
        sorted_students = sorted(students, key=lambda s: (s.group_name, s.full_name))
        for i, s in enumerate(sorted_students, 1):
            reg = s.registered_at.strftime("%d.%m.%Y") if s.registered_at else "—"
            ws.append([i, s.full_name, s.group_name, s.mars_id,
                       s.phone_number or "—",
                       s.telegram_username or str(s.user_id), reg])

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"oquvchilar_{date_str}.xlsx"
        file = BufferedInputFile(buf.read(), filename=filename)

        await bot.send_document(
            chat_id=cb.from_user.id,
            document=file,
            caption=f"📥 O'quvchilar ro'yxati | {len(students)} ta | {date_str}",
        )
    except ImportError:
        await cb.message.answer("❌ openpyxl o'rnatilmagan. pip install openpyxl")
    except Exception as e:
        await cb.message.answer(f"❌ Xato: <code>{e}</code>")


# ════════════════════════════════════════════════════════════════════════════════
# BROADCAST
# ════════════════════════════════════════════════════════════════════════════════

class BroadcastFSM(StatesGroup):
    waiting_target  = State()
    waiting_message = State()


@router.callback_query(F.data == "admin:broadcast")
async def admin_broadcast_start(cb: CallbackQuery, state: FSMContext) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    await state.set_state(BroadcastFSM.waiting_target)
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📢 Barcha o'quvchilar", callback_data="bc:target:all"))
    for g in MARS_GROUPS:
        builder.row(InlineKeyboardButton(text=f"📚 {g}", callback_data=f"bc:target:{g}"))
    builder.row(InlineKeyboardButton(text="◀️ Bekor", callback_data="admin:panel"))

    try:
        await cb.message.edit_text("📢 <b>Broadcast</b>\n\nKimga yuboramiz?", reply_markup=builder.as_markup())
    except TelegramBadRequest:
        await cb.message.answer("📢 <b>Broadcast</b>\n\nKimga yuboramiz?", reply_markup=builder.as_markup())
    await cb.answer()


@router.callback_query(F.data.startswith("bc:target:"), StateFilter(BroadcastFSM.waiting_target))
async def admin_broadcast_target(cb: CallbackQuery, state: FSMContext) -> None:
    target = cb.data.split(":", 2)[2]
    await state.update_data(target=target)
    await state.set_state(BroadcastFSM.waiting_message)
    label = "barcha o'quvchilarga" if target == "all" else f"{target} guruhiga"
    await cb.message.edit_text(
        f"📢 Xabar <b>{label}</b> yuboriladi.\n\n"
        "Xabarni yuboring (matn, rasm, video, fayl — istalgan):\n\n"
        "<i>Bekor qilish: /start</i>",
    )
    await cb.answer()


@router.message(StateFilter(BroadcastFSM.waiting_message))
async def admin_broadcast_send(message: Message, state: FSMContext, db: DatabaseService, bot: Bot) -> None:
    data   = await state.get_data()
    target = data.get("target", "all")
    await state.clear()

    if target == "all":
        students = await db.get_all_students()
    else:
        students = await db.get_students_by_group(target)

    ok, fail = 0, 0
    for s in students:
        try:
            await bot.copy_message(s.user_id, message.chat.id, message.message_id)
            ok += 1
        except Exception:
            fail += 1

    label = "Barcha o'quvchilar" if target == "all" else target
    await message.answer(
        f"✅ <b>Broadcast yuborildi!</b>\n\n"
        f"👥 Maqsad: <b>{label}</b>\n"
        f"✅ Yetdi: <b>{ok}</b> ta | ❌ Yetmadi: <b>{fail}</b> ta",
        reply_markup=kb_back_to_panel(),
    )


# ════════════════════════════════════════════════════════════════════════════════
# YANGI O'QUVCHI QO'SHISH
# ════════════════════════════════════════════════════════════════════════════════

class AddCredentialFSM(StatesGroup):
    waiting_group    = State()
    waiting_name     = State()
    waiting_mars_id  = State()
    waiting_password = State()


@router.callback_query(F.data == "admin:add_student_cred")
async def admin_add_cred_start(cb: CallbackQuery, state: FSMContext) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return
    await state.set_state(AddCredentialFSM.waiting_group)
    try:
        await cb.message.edit_text(
            "➕ <b>Yangi o'quvchi qo'shish</b>\n\nGuruhni tanlang:",
            reply_markup=kb_hw_groups(MARS_GROUPS, prefix="addcred"),
        )
    except TelegramBadRequest:
        await cb.message.answer(
            "➕ <b>Yangi o'quvchi qo'shish</b>\n\nGuruhni tanlang:",
            reply_markup=kb_hw_groups(MARS_GROUPS, prefix="addcred"),
        )
    await cb.answer()


@router.callback_query(F.data.startswith("addcred:group:"), StateFilter(AddCredentialFSM.waiting_group))
async def admin_add_cred_group(cb: CallbackQuery, state: FSMContext) -> None:
    group = cb.data.split(":", 2)[2]
    await state.update_data(group=group)
    await state.set_state(AddCredentialFSM.waiting_name)
    await cb.message.edit_text(
        f"📚 Guruh: <b>{group}</b>\n\nO'quvchining <b>ism familyasini</b> kiriting:"
    )
    await cb.answer()


@router.message(StateFilter(AddCredentialFSM.waiting_name))
async def admin_add_cred_name(message: Message, state: FSMContext) -> None:
    await state.update_data(name=message.text.strip())
    await state.set_state(AddCredentialFSM.waiting_mars_id)
    await message.answer("🔑 <b>Mars Space ID</b> raqamini kiriting:")


@router.message(StateFilter(AddCredentialFSM.waiting_mars_id))
async def admin_add_cred_mars_id(message: Message, state: FSMContext) -> None:
    mars_id = message.text.strip()
    if not mars_id.isdigit():
        await message.answer("❌ ID faqat raqam bo'lishi kerak. Qayta kiriting:")
        return
    await state.update_data(mars_id=mars_id)
    await state.set_state(AddCredentialFSM.waiting_password)
    await message.answer("🔐 <b>Parolni</b> kiriting:")


@router.message(StateFilter(AddCredentialFSM.waiting_password))
async def admin_add_cred_password(message: Message, state: FSMContext, db: DatabaseService) -> None:
    data     = await state.get_data()
    group    = data.get("group", "")
    name     = data.get("name", "")
    mars_id  = data.get("mars_id", "")
    password = message.text.strip()
    await state.clear()

    await db.add_student_credential(mars_id, name, password, group)
    await message.answer(
        f"✅ <b>O'quvchi qo'shildi!</b>\n\n"
        f"👤 Ism: <b>{name}</b>\n"
        f"📚 Guruh: <b>{group}</b>\n"
        f"🆔 Mars ID: <code>{mars_id}</code>\n\n"
        f"<i>O'quvchi endi /start bossa ro'yxatdan o'ta oladi.</i>",
        reply_markup=kb_back_to_panel(),
    )


# ════════════════════════════════════════════════════════════════════════════════
# ESLATMA VAQTINI O'ZGARTIRISH
# ════════════════════════════════════════════════════════════════════════════════

class TimeFSM(StatesGroup):
    waiting_time = State()


@router.callback_query(F.data == "admin:set_time")
async def admin_set_time_start(cb: CallbackQuery, state: FSMContext, db: DatabaseService) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    cur_h = await db.get_setting("SEND_HOUR", str(SEND_HOUR))
    cur_m = await db.get_setting("SEND_MINUTE", str(SEND_MINUTE))
    await state.set_state(TimeFSM.waiting_time)
    try:
        await cb.message.edit_text(
            f"⏰ <b>Eslatma vaqtini o'zgartirish</b>\n\n"
            f"Hozirgi vaqt: <b>{int(cur_h):02d}:{int(cur_m):02d}</b>\n\n"
            f"Yangi vaqtni kiriting (masalan: <code>19:30</code> yoki <code>20:00</code>):\n\n"
            f"<i>Bekor qilish: /start</i>",
        )
    except TelegramBadRequest:
        await cb.message.answer(
            f"⏰ Yangi vaqtni kiriting (HH:MM):\n<i>Bekor qilish: /start</i>",
        )
    await cb.answer()


@router.message(StateFilter(TimeFSM.waiting_time))
async def admin_set_time_save(message: Message, state: FSMContext, db: DatabaseService) -> None:
    text = message.text.strip()
    try:
        h_str, m_str = text.split(":")
        hour   = int(h_str)
        minute = int(m_str)
        assert 0 <= hour <= 23 and 0 <= minute <= 59
    except Exception:
        await message.answer("❌ Format noto'g'ri! Masalan: <code>20:00</code> yoki <code>19:30</code>")
        return

    await state.clear()
    await db.set_setting("SEND_HOUR", str(hour))
    await db.set_setting("SEND_MINUTE", str(minute))

    # Schedulerni qayta sozlaymiz
    from scheduler import reschedule_reminder
    reschedule_reminder(hour, minute)

    await message.answer(
        f"✅ Eslatma vaqti o'zgartirildi: <b>{hour:02d}:{minute:02d}</b>\n\n"
        f"<i>Botni qayta ishga tushirmasdan ishlaydi.</i>",
        reply_markup=kb_back_to_panel(),
    )


# ════════════════════════════════════════════════════════════════════════════════
# FAOLLIK TEKSHIRISH
# ════════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data == "admin:check_activity")
async def admin_check_activity(cb: CallbackQuery, db: DatabaseService) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    inactive = await db.get_inactive_students(days=7)

    if not inactive:
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="◀️ Panel", callback_data="admin:panel"))
        try:
            await cb.message.edit_text(
                "✅ <b>Barcha o'quvchilar faol!</b>\n\n"
                "So'nggi 7 kun ichida hamma botga kirgan.",
                reply_markup=builder.as_markup(),
            )
        except TelegramBadRequest:
            await cb.message.answer("✅ Barcha o'quvchilar faol!", reply_markup=builder.as_markup())
        await cb.answer()
        return

    builder = InlineKeyboardBuilder()
    for s in inactive[:20]:  # max 20 ta
        tg = s.telegram_username or f"ID:{s.user_id}"
        last = s.last_active.strftime("%d.%m") if s.last_active else "Hech qachon"
        builder.row(InlineKeyboardButton(
            text=f"👤 {s.full_name} | {s.group_name} | {last}",
            callback_data=f"admin:student_detail:{s.user_id}",
        ))
    builder.row(InlineKeyboardButton(text="◀️ Panel", callback_data="admin:panel"))

    text = (
        f"🔍 <b>Nofaol o'quvchilar (7+ kun)</b>\n\n"
        f"Jami: <b>{len(inactive)}</b> ta\n"
        f"<i>Ismiga bosib xabar yuborishingiz mumkin</i>"
    )
    try:
        await cb.message.edit_text(text, reply_markup=builder.as_markup())
    except TelegramBadRequest:
        await cb.message.answer(text, reply_markup=builder.as_markup())
    await cb.answer()


# ════════════════════════════════════════════════════════════════════════════════
# O'QUVCHI: UY VAZIFASI TARIXI
# ════════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data == "student:hw_history")
async def student_hw_history(cb: CallbackQuery, db: DatabaseService, bot: Bot) -> None:
    student = await db.get_student(cb.from_user.id)
    if not student:
        await cb.answer("❌ Avval ro'yxatdan o'ting!", show_alert=True); return

    await db.update_last_active(cb.from_user.id)
    history = await db.get_homework_history(student.group_name, limit=5)
    if not history:
        await cb.answer("📭 Hozircha uy vazifasi tarixi yo'q.", show_alert=True); return

    await cb.message.answer(
        f"📜 <b>{student.group_name} — Oxirgi {len(history)} ta uy vazifasi:</b>"
    )
    for hw in history:
        date_str = hw.sent_at.strftime("%d.%m.%Y %H:%M")
        try:
            await bot.copy_message(
                chat_id=cb.from_user.id,
                from_chat_id=hw.from_chat_id,
                message_id=hw.message_id,
            )
        except Exception:
            await cb.message.answer(f"⚠️ {date_str} — yuklashda xato")
    await cb.answer()


# ════════════════════════════════════════════════════════════════════════════════
# GURUHGA UMUMIY XABAR YUBORISH (O'QUVCHILAR / OTA-ONALAR)
# ════════════════════════════════════════════════════════════════════════════════

class GroupMsgFSM(StatesGroup):
    waiting_content = State()


@router.callback_query(F.data.startswith("admin:grp_msg:"))
async def admin_grp_msg_start(cb: CallbackQuery, state: FSMContext) -> None:
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("❌", show_alert=True); return

    target = cb.data.split(":")[2]   # "student" | "parent"
    label  = "O'quvchilar guruhlariga" if target == "student" else "Ota-onalar guruhlariga"

    await state.update_data(grp_target=target)
    await state.set_state(GroupMsgFSM.waiting_content)
    try:
        await cb.message.edit_text(
            f"📢 <b>{label}</b> xabar yuboriladi.\n\n"
            f"Xabarni yuboring (matn, rasm, video, fayl — istalgan):\n\n"
            f"<i>Bekor qilish: /start</i>",
        )
    except TelegramBadRequest:
        await cb.message.answer(
            f"📢 <b>{label}</b> — xabarni yuboring:\n<i>Bekor: /start</i>"
        )
    await cb.answer()


@router.message(StateFilter(GroupMsgFSM.waiting_content))
async def admin_grp_msg_send(
    message: Message, state: FSMContext, db: DatabaseService, bot: Bot
) -> None:
    data   = await state.get_data()
    target = data.get("grp_target", "student")
    await state.clear()

    audience = AudienceType.STUDENT if target == "student" else AudienceType.PARENT
    groups   = await db.get_all_groups()
    targets  = [g for g in groups if g.audience == audience and g.is_active]

    label = "O'quvchilar" if target == "student" else "Ota-onalar"
    ok, fail = 0, 0

    for g in targets:
        try:
            await bot.copy_message(
                chat_id=g.chat_id,
                from_chat_id=message.chat.id,
                message_id=message.message_id,
            )
            ok += 1
        except Exception as e:
            fail += 1
            logger.warning(f"Guruhga xabar yuborib bo'lmadi '{g.name}': {e}")

    await message.answer(
        f"✅ <b>{label} guruhlariga xabar yuborildi!</b>\n\n"
        f"✅ Yetdi: <b>{ok}</b> ta guruh\n"
        f"❌ Yetmadi: <b>{fail}</b> ta guruh",
        reply_markup=kb_back_to_panel(),
    )
